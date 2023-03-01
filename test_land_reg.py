from playwright.sync_api import Page, expect
import  openpyxl
from pathlib import Path




def test_land_use_act_and_scrape_the_text(page: Page):
    data = []
    titles = []
    dataframe = openpyxl.load_workbook(Path().joinpath('land_use_reg.xlsx'))
    dataframe1 = dataframe.active
    # cell = dataframe1.cell(row=2, column=1)
    # print('hello')
    # print("result: \n"+cell.value)
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            data.append(col[row].value)

    # Get to the search Page
    page.goto("https://find-and-update.company-information.service.gov.uk/")
    page.get_by_role("button", name="Reject analytics cookies").click()
    page.get_by_role("button", name="Hide this message").click()


    # page.goto("https://find-and-update.company-information.service.gov.uk/")
    # page.get_by_role("button", name="Reject analytics cookies").click()
    # page.get_by_role("button", name="Hide this message").click()
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").fill("ELDRITCH")
    # page.get_by_role("button", name="Search").click()
    # page.get_by_role("link", name="ELDRITCH LLP").click()
    # page.get_by_text("Company number OC385139").click()
    # page.goto("https://find-and-update.company-information.service.gov.uk/")
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("QUEENTON LIMITED")
    # page.get_by_role("button", name="Search").click()
    # page.goto("https://find-and-update.company-information.service.gov.uk/")
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("DUNBAR VENTURES CORP.")
    # page.get_by_role("button", name="Search").click()
    # page.get_by_role("link", name="DUNBAR VENTURES CORP.").click()
    # page.get_by_text("Company number OE014556").click()
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("GESTRUST SA")
    # page.get_by_role("button", name="Search").click()
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("BARMOUTH LIMITED")
    # page.get_by_role("button", name="Search").click()
    # page.get_by_role("link", name="BARMOUTH LIMITED").click()
    # page.get_by_text("Company number OE005420").click()
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("DREEMGARROW LIMITED")
    # page.get_by_role("button", name="Search").click()
    # page.get_by_role("link", name="DREEMGARROW LIMITED").click()
    # page.get_by_text("Company number OE007321").click()
    # page.get_by_placeholder("Start here...").click()
    # page.get_by_placeholder("Start here...").press("Control+a")
    # page.get_by_placeholder("Start here...").fill("KAA CONSULTING S.A. BVI")
    # page.get_by_role("button", name="Search").click()
    # page.goto("https://find-and-update.company-information.service.gov.uk/")



    for item in data:
        page.get_by_placeholder("Start here...").click()
        page.get_by_placeholder("Start here...").fill(item)
        page.get_by_role("button", name="Search").click()
        try:
            link = page.get_by_role("link", name=item[:8])
            if(link.is_visible()):
                link.click()
                title = page.title()
                titles.append(title)
            page.go_back()
        except NameError:
            print('Nothing of such found')
            continue
    print(titles)
    page.close()